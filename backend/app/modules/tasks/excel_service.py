"""
SmartTask360 — Excel service for task import/export
"""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.tasks.excel_schemas import ImportErrorDetail, ImportResult
from app.modules.tasks.models import Task
from app.modules.tasks.schemas import TaskCreate
from app.modules.tasks.service import TaskService
from app.modules.users.models import User


# Column definitions for export/import
EXPORT_COLUMNS = [
    ("id", "ID", 36),
    ("title", "Название", 50),
    ("description", "Описание", 60),
    ("status", "Статус", 15),
    ("priority", "Приоритет", 12),
    ("author_email", "Автор (email)", 30),
    ("creator_email", "Постановщик (email)", 30),
    ("assignee_email", "Исполнитель (email)", 30),
    ("parent_id", "Родительская задача (ID)", 36),
    ("depth", "Уровень вложенности", 10),
    ("department_id", "Отдел (ID)", 36),
    ("project_id", "Проект (ID)", 36),
    ("due_date", "Срок", 20),
    ("started_at", "Начата", 20),
    ("completed_at", "Завершена", 20),
    ("is_milestone", "Веха", 8),
    ("estimated_hours", "Оценка (часы)", 12),
    ("actual_hours", "Факт (часы)", 12),
    ("acceptance_deadline", "Дедлайн принятия", 20),
    ("rejection_reason", "Причина отклонения", 20),
    ("rejection_comment", "Комментарий отклонения", 40),
    ("smart_is_valid", "SMART валидна", 12),
    ("created_at", "Создана", 20),
    ("updated_at", "Обновлена", 20),
]

# Columns that can be imported (writable)
IMPORT_COLUMNS = [
    "title",
    "description",
    "status",
    "priority",
    "creator_email",
    "assignee_email",
    "parent_id",
    "department_id",
    "project_id",
    "due_date",
    "is_milestone",
    "estimated_hours",
    "acceptance_deadline",
]

# Valid enum values
VALID_STATUSES = ["new", "assigned", "in_progress", "in_review", "on_hold", "done", "cancelled", "draft"]
VALID_PRIORITIES = ["low", "medium", "high", "critical"]

# Russian translations for statuses and priorities
STATUS_RU = {
    "new": "Новая",
    "assigned": "Назначена",
    "in_progress": "В работе",
    "in_review": "На проверке",
    "on_hold": "На паузе",
    "done": "Готово",
    "cancelled": "Отменена",
    "draft": "Черновик",
}

PRIORITY_RU = {
    "low": "Низкий",
    "medium": "Средний",
    "high": "Высокий",
    "critical": "Критический",
}


class ExcelService:
    """Service for Excel import/export of tasks"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._user_cache: dict[str, UUID] = {}  # email -> user_id
        self._user_email_cache: dict[UUID, str] = {}  # user_id -> email

    async def _build_user_caches(self) -> None:
        """Build email <-> user_id mapping caches"""
        result = await self.db.execute(select(User.id, User.email))
        for user_id, email in result.all():
            self._user_cache[email.lower()] = user_id
            self._user_email_cache[user_id] = email

    def _get_user_email(self, user_id: UUID | None) -> str:
        """Get user email by ID"""
        if not user_id:
            return ""
        return self._user_email_cache.get(user_id, "")

    def _get_user_id_by_email(self, email: str | None) -> UUID | None:
        """Get user ID by email"""
        if not email:
            return None
        return self._user_cache.get(email.lower().strip())

    async def export_tasks(
        self,
        status: str | None = None,
        priority: str | None = None,
        search: str | None = None,
        project_id: UUID | None = None,
        department_id: UUID | None = None,
    ) -> bytes:
        """Export tasks to Excel file"""
        await self._build_user_caches()

        # Get tasks with filters
        task_service = TaskService(self.db)
        tasks = await task_service.get_all(
            skip=0,
            limit=10000,  # Large limit for export
            include_deleted=False,
            status=status,
            priority=priority,
            search=search,
            project_id=project_id,
        )

        # Filter by department_id if provided
        if department_id:
            tasks = [t for t in tasks if t.department_id == department_id]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (field, header, width) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, task in enumerate(tasks, start=2):
            for col_idx, (field, _, _) in enumerate(EXPORT_COLUMNS, start=1):
                value = self._get_task_field_value(task, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_task_field_value(self, task: Task, field: str):
        """Get task field value for export"""
        if field == "id":
            return str(task.id) if task.id else ""
        elif field == "author_email":
            return self._get_user_email(task.author_id)
        elif field == "creator_email":
            return self._get_user_email(task.creator_id)
        elif field == "assignee_email":
            return self._get_user_email(task.assignee_id)
        elif field in ("parent_id", "department_id", "project_id"):
            value = getattr(task, field, None)
            return str(value) if value else ""
        elif field in ("due_date", "started_at", "completed_at", "created_at", "updated_at", "acceptance_deadline"):
            value = getattr(task, field, None)
            if value:
                return value.strftime("%Y-%m-%d %H:%M") if isinstance(value, datetime) else str(value)
            return ""
        elif field in ("is_milestone", "smart_is_valid"):
            value = getattr(task, field, None)
            if value is None:
                return ""
            return "Да" if value else "Нет"
        elif field in ("estimated_hours", "actual_hours"):
            value = getattr(task, field, None)
            return float(value) if value else ""
        elif field == "status":
            value = getattr(task, field, "")
            return STATUS_RU.get(value, value)
        elif field == "priority":
            value = getattr(task, field, "")
            return PRIORITY_RU.get(value, value)
        else:
            return getattr(task, field, "") or ""

    def generate_template(self) -> bytes:
        """Generate Excel template for import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Шаблон импорта"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        optional_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Template columns (subset for import)
        template_columns = [
            ("title", "Название *", 50, True),
            ("description", "Описание", 60, False),
            ("status", "Статус", 15, False),
            ("priority", "Приоритет", 12, False),
            ("assignee_email", "Исполнитель (email)", 30, False),
            ("due_date", "Срок (ГГГГ-ММ-ДД)", 20, False),
            ("estimated_hours", "Оценка (часы)", 12, False),
            ("is_milestone", "Веха (Да/Нет)", 12, False),
            ("project_id", "Проект (ID)", 36, False),
            ("department_id", "Отдел (ID)", 36, False),
        ]

        # Write headers
        for col_idx, (field, header, width, required) in enumerate(template_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = required_fill if required else optional_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Add example row
        example_values = [
            "Название задачи",
            "Описание задачи",
            "Новая",
            "Средний",
            "user@example.com",
            "2025-01-15",
            "8",
            "Нет",
            "",
            "",
        ]
        for col_idx, value in enumerate(example_values, start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top")

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add instructions sheet
        ws_info = wb.create_sheet("Инструкция")
        instructions = [
            ("Инструкция по импорту задач", None),
            ("", None),
            ("Обязательные поля:", None),
            ("- Название: Название задачи (обязательно)", None),
            ("", None),
            ("Необязательные поля:", None),
            ("- Описание: Подробное описание задачи", None),
            ("- Статус: Новая, Назначена, В работе, На проверке, На паузе, Готово, Отменена", None),
            ("- Приоритет: Низкий, Средний, Высокий, Критический", None),
            ("- Исполнитель: Email пользователя (должен существовать в системе)", None),
            ("- Срок: Дата в формате ГГГГ-ММ-ДД (например: 2025-01-15)", None),
            ("- Оценка: Количество часов (число)", None),
            ("- Веха: Да или Нет", None),
            ("- Проект ID: UUID проекта (для привязки к проекту)", None),
            ("- Отдел ID: UUID отдела (для привязки к отделу)", None),
        ]
        for row_idx, (text, _) in enumerate(instructions, start=1):
            cell = ws_info.cell(row=row_idx, column=1, value=text)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif text.startswith("-"):
                cell.font = Font(size=11)
        ws_info.column_dimensions["A"].width = 80

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def import_tasks(self, file_data: bytes, user_id: UUID) -> ImportResult:
        """Import tasks from Excel file"""
        await self._build_user_caches()

        errors: list[ImportErrorDetail] = []
        imported = 0
        skipped = 0
        total_rows = 0

        try:
            wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        except Exception as e:
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[
                    ImportErrorDetail(
                        row=0,
                        field="file",
                        message=f"Не удалось прочитать файл: {str(e)}",
                        value=None,
                    )
                ],
            )

        ws = wb.active
        if not ws:
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[
                    ImportErrorDetail(row=0, field="file", message="Файл не содержит данных", value=None)
                ],
            )

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(self._normalize_header(cell.value))

        # Map header names to column indices
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        # Process data rows
        task_service = TaskService(self.db)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip empty rows

            total_rows += 1
            row_errors: list[ImportErrorDetail] = []

            try:
                task_data = self._parse_row(row, header_map, row_idx, row_errors)

                if row_errors:
                    errors.extend(row_errors)
                    skipped += 1
                    continue

                if not task_data:
                    skipped += 1
                    continue

                # Create task
                await task_service.create(task_data, user_id)
                imported += 1

            except Exception as e:
                errors.append(
                    ImportErrorDetail(
                        row=row_idx,
                        field="general",
                        message=f"Ошибка создания задачи: {str(e)}",
                        value=None,
                    )
                )
                skipped += 1

        wb.close()

        return ImportResult(
            success=len(errors) == 0,
            total_rows=total_rows,
            imported=imported,
            skipped=skipped,
            errors=errors,
        )

    def _normalize_header(self, header: str | None) -> str | None:
        """Normalize header name to field name"""
        if not header:
            return None

        header = str(header).strip().lower()

        # Map Russian headers to field names
        mapping = {
            "название": "title",
            "название *": "title",
            "описание": "description",
            "статус": "status",
            "приоритет": "priority",
            "исполнитель (email)": "assignee_email",
            "исполнитель": "assignee_email",
            "срок": "due_date",
            "срок (гггг-мм-дд)": "due_date",
            "оценка (часы)": "estimated_hours",
            "оценка": "estimated_hours",
            "веха": "is_milestone",
            "веха (да/нет)": "is_milestone",
            "проект (id)": "project_id",
            "проект": "project_id",
            "отдел (id)": "department_id",
            "отдел": "department_id",
            "постановщик (email)": "creator_email",
            "постановщик": "creator_email",
            "родительская задача (id)": "parent_id",
        }

        return mapping.get(header, header)

    def _parse_row(
        self,
        row: tuple,
        header_map: dict[str, int],
        row_idx: int,
        errors: list[ImportErrorDetail],
    ) -> TaskCreate | None:
        """Parse row data into TaskCreate schema"""
        from app.core.types import TaskPriority, TaskStatus

        def get_cell(field: str) -> str | None:
            idx = header_map.get(field)
            if idx is None or idx >= len(row):
                return None
            value = row[idx]
            if value is None:
                return None
            return str(value).strip()

        # Required field: title
        title = get_cell("title")
        if not title:
            errors.append(
                ImportErrorDetail(row=row_idx, field="title", message="Название обязательно", value=None)
            )
            return None

        # Parse status
        status_str = get_cell("status")
        status = TaskStatus.NEW
        if status_str:
            status_key = self._parse_status(status_str)
            if status_key:
                try:
                    status = TaskStatus(status_key)
                except ValueError:
                    errors.append(
                        ImportErrorDetail(
                            row=row_idx,
                            field="status",
                            message=f"Неизвестный статус: {status_str}",
                            value=status_str,
                        )
                    )

        # Parse priority
        priority_str = get_cell("priority")
        priority = TaskPriority.MEDIUM
        if priority_str:
            priority_key = self._parse_priority(priority_str)
            if priority_key:
                try:
                    priority = TaskPriority(priority_key)
                except ValueError:
                    errors.append(
                        ImportErrorDetail(
                            row=row_idx,
                            field="priority",
                            message=f"Неизвестный приоритет: {priority_str}",
                            value=priority_str,
                        )
                    )

        # Parse assignee
        assignee_email = get_cell("assignee_email")
        assignee_id = None
        if assignee_email:
            assignee_id = self._get_user_id_by_email(assignee_email)
            if not assignee_id:
                errors.append(
                    ImportErrorDetail(
                        row=row_idx,
                        field="assignee_email",
                        message=f"Пользователь не найден: {assignee_email}",
                        value=assignee_email,
                    )
                )

        # Parse creator
        creator_email = get_cell("creator_email")
        creator_id = None
        if creator_email:
            creator_id = self._get_user_id_by_email(creator_email)
            if not creator_id:
                errors.append(
                    ImportErrorDetail(
                        row=row_idx,
                        field="creator_email",
                        message=f"Пользователь не найден: {creator_email}",
                        value=creator_email,
                    )
                )

        # Parse due_date
        due_date = None
        due_date_str = get_cell("due_date")
        if due_date_str:
            due_date = self._parse_date(due_date_str)
            if not due_date:
                errors.append(
                    ImportErrorDetail(
                        row=row_idx,
                        field="due_date",
                        message=f"Неверный формат даты: {due_date_str}",
                        value=due_date_str,
                    )
                )

        # Parse estimated_hours
        estimated_hours = None
        estimated_str = get_cell("estimated_hours")
        if estimated_str:
            try:
                estimated_hours = Decimal(estimated_str.replace(",", "."))
            except (ValueError, TypeError):
                errors.append(
                    ImportErrorDetail(
                        row=row_idx,
                        field="estimated_hours",
                        message=f"Неверное число: {estimated_str}",
                        value=estimated_str,
                    )
                )

        # Parse is_milestone
        is_milestone = False
        milestone_str = get_cell("is_milestone")
        if milestone_str:
            is_milestone = milestone_str.lower() in ("да", "yes", "true", "1")

        # Parse UUIDs
        project_id = self._parse_uuid(get_cell("project_id"))
        department_id = self._parse_uuid(get_cell("department_id"))
        parent_id = self._parse_uuid(get_cell("parent_id"))

        # If there were errors, don't create task but continue parsing others
        if errors:
            return None

        return TaskCreate(
            title=title,
            description=get_cell("description"),
            status=status,
            priority=priority,
            creator_id=creator_id,
            assignee_id=assignee_id,
            parent_id=parent_id,
            department_id=department_id,
            project_id=project_id,
            due_date=due_date,
            is_milestone=is_milestone,
            estimated_hours=estimated_hours,
        )

    def _parse_status(self, status_str: str) -> str | None:
        """Parse status from Russian or English"""
        status_str = status_str.strip().lower()

        # Reverse mapping from Russian
        ru_to_key = {v.lower(): k for k, v in STATUS_RU.items()}

        if status_str in VALID_STATUSES:
            return status_str
        elif status_str in ru_to_key:
            return ru_to_key[status_str]
        return None

    def _parse_priority(self, priority_str: str) -> str | None:
        """Parse priority from Russian or English"""
        priority_str = priority_str.strip().lower()

        # Reverse mapping from Russian
        ru_to_key = {v.lower(): k for k, v in PRIORITY_RU.items()}

        if priority_str in VALID_PRIORITIES:
            return priority_str
        elif priority_str in ru_to_key:
            return ru_to_key[priority_str]
        return None

    def _parse_date(self, date_str: str) -> datetime | None:
        """Parse date from various formats"""
        formats = [
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y",
            "%d.%m.%Y %H:%M",
            "%d/%m/%Y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue
        return None

    def _parse_uuid(self, uuid_str: str | None) -> UUID | None:
        """Parse UUID from string"""
        if not uuid_str:
            return None
        try:
            return UUID(uuid_str.strip())
        except (ValueError, TypeError):
            return None
